import openpyxl
import MySQLdb
conn = MySQLdb.connect(host = "127.0.0.1",user = "root",password = "123456",database = "bot",charset = "utf8")

# 处理ID为空的情况
# path = '/Users/yangchiher/Desktop/文本机器人id整合5.9.xlsx'
# wb = openpyxl.load_workbook(path)
# sheet1 = wb['Sheet1']
# ad = sheet1['AD']
# for i in range(len(ad)):
#     cell = ad[i]
#     if(cell.value):
#         print(i, str(cell.value))
#         categoryId = cell.value
#     else:
#         cell.value = categoryId
# wb.save(path)

# path = '/Users/yangchiher/Desktop/专业知识点-分类详情.xlsx'
# wb = openpyxl.load_workbook(path)
# sheet1 = wb['Sheet2']
# id = sheet1['A']
# name = sheet1['B']
# parentId = sheet1['C']
# count = 0
# for i in range(len(id)):
#     id_ = id[i]
#     name_ = name[i]
#     parentId_ = parentId[i]
#     #print(id_.value,name_.value,parentId_.value)
#
#     faqSql = "select * from categories where name = " + "'"+name_.value+"'"
#
#     cursor_ = conn.cursor()
#     cursor_.execute(faqSql)
#     result_ = cursor_.fetchall()
#     cursor_.close()
#
#     #print(result_)
#     if len(result_)==0:
#         print(count,'###',name_.value)
#         count += 1


#wb.save(path)


path = '/Users/yangchiher/Desktop/bot.xlsx'
wb = openpyxl.load_workbook(path)
sheet1 = wb['Sheet1']
name1 = sheet1['A']
name2 = sheet1['B']
name3 = sheet1['C']
count = 0
for i in range(len(name1)):
    name1_ = name1[i]
    name2_ = name2[i]
    name3_ = name3[i]

    categoryName = name3_.value

    if categoryName and i > 0:
        faqSql = "select * from categories where name = " + "'"+categoryName+"'"

        cursor_ = conn.cursor()
        cursor_.execute(faqSql)
        result_ = cursor_.fetchall()
        cursor_.close()

        if len(result_)==0:
            #print(count,'###',categoryName)
            count += 1
        else:
            print(count, '正常', name1_.value)


